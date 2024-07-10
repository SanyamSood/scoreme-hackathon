import pdfplumber
from openpyxl import Workbook

# Load the PDF
pdf_path = 'test3.pdf'
pdf = pdfplumber.open(pdf_path)

# Initialize Excel workbook and sheet
wb = Workbook()
sheet = wb.active
sheet.title = 'Extracted Table'

# Write column headings
sheet.cell(row=1, column=1, value="Date")
sheet.cell(row=1, column=2, value="Transaction Info")
sheet.cell(row=1, column=3, value="Amount")
#sheet.cell(row=1, column=4, value="Balance")

# Flag to start capturing table data
capture = False
row_index = 2  # Start writing from row 2 in Excel (row 1 is for headings)

# Iterate through pages
for page in pdf.pages:
    text = page.extract_text()
    lines = text.split('\n')
    
    # Flag to check for the start of table data
    start_index = None
    
    # Find the starting point of the table
    for i, line in enumerate(lines):
        if "Statement of account" in line:
            start_index = i + 1  # Start capturing after this line
            break
    
    if start_index is not None:
        capture = True
    
    if capture:
        # Process lines from the identified start point
        current_row = []
        for line in lines[start_index:]:
            if "Dr" in line:
                # End of current row, process and write to Excel
                current_row.append(line.strip())
                # Combine lines to form complete transaction entry
                if len(current_row) > 1:
                    # Join lines to form complete transaction entry
                    transaction_entry = ' '.join(current_row)
                    
                    # Split the transaction entry into columns based on expected format
                    date = current_row[0][:11].strip()
                    transaction_info = current_row[0][11:].strip()
                    
                    # Find the index where "Dr" occurs
                    dr_index = current_row[1].find("Dr")
                    
                    # Split amount and remaining balance from each line
                    if dr_index != -1:
                        amount_line = current_row[1][:dr_index].strip()
                        remaining_balance_line = current_row[1][dr_index:].strip().replace("Dr", "").strip()
                    else:
                        amount_line = current_row[1].strip()
                        remaining_balance_line = ""
                    
                    # Write data to Excel
                    sheet.cell(row=row_index, column=1, value=date)
                    sheet.cell(row=row_index, column=2, value=transaction_info)
                    sheet.cell(row=row_index, column=3, value=amount_line)
                    sheet.cell(row=row_index, column=4, value=remaining_balance_line)
                    
                    row_index += 1  # Move to the next row
                    current_row = []  # Reset for the next row
            else:
                # Append line to current row
                current_row.append(line.strip())
    
    # You might need additional logic to handle page breaks or other irregularities
    
# Save Excel workbook
excel_output_path = 'extracted.xlsx'
wb.save(excel_output_path)

# Close PDF
pdf.close()

